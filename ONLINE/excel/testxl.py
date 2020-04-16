import openpyxl as xl

# workbook
#   worksheet
#     rows/columns
#       cells (address of column letter and row number)

wb = xl.load_workbook("test.xlsx")
# print(wb.worksheets)
# for ws in wb.worksheets:
#     print(ws.title)

ws = wb.active         # whatever sheet was active when last saved
ws = wb.worksheets[0]  # specific sheet from the list of all sheets

for row in ws.rows:    # iterate over all cells
    for cell in row:
        print(cell.value)


print(ws['A2'].value)  # access cells in a "random access" kind of way

for row in ws['A2:B5']:
    for cell in row:
        print(cell.value)

cell = ws['A1']
print(dir(cell))
print(cell.col_idx)
print(cell.column)
print(cell.column_letter)
print(cell.row)

from openpyxl.utils import get_column_letter as gcl

print(gcl(5))



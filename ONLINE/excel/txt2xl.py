from pprint import pprint as pp
import openpyxl as xl
import os

grades_dict = {}
# {'alan.alda': [75, 82, 73],
#  'benny.barker': [85, 75, 88],
#  'chloe.carter': [100, 100, 100],
#  'demarcus.dalton': [95, 91, 94],
#  'elena.etcher': [65, 81, 91],
#  'farad.fami': [80, 73, 85],
#  'gizelle.garner': [92, 94, 100],
#  'helena.homer': [88, 91, 100],
#  'issac.imogen': [58, 71, 82],
#  'jessica.jordan': [88, 84, 91],
#  'katya.kruschev': [74, 72, 81],
#  'latisha.love': [98, 95, 78],
#  'marco.massey': [83, 78, 92]}

grade_count = 0

for filename in os.listdir():
    # print(filename)
    if filename.startswith("grades") and filename.endswith(".txt"):
        with open(filename) as file:
            grade_count += 1

            for line in file:
                line = line.strip()
                vals = line.split(',')  # vals[0] is name and vals[1] is grade
                name = vals[0]
                grade = vals[1]     # is a string
                grade = int(grade)  # now is an int
                
                if name in grades_dict:
                    # if it's already there, add it to the list of grades
                    grades = grades_dict[name]  # grades will be a list
                    grades.append(grade)
                else:
                    grades_dict[name] = [grade]

# pp(grades_dict)

wb = xl.Workbook()    # create a NEW workbook

wb.remove(wb.active)

for num in range(grade_count):  # (0,1,2)
    ws = wb.create_sheet(f"Grades for Exam {num + 1}")    # create a NEW worksheet
    row_count = 0
    for name in grades_dict:        # gets the key
        grades = grades_dict[name]  # gets the list
        row_count += 1
        ws[f"A{row_count}"].value = name
        ws[f"B{row_count}"].value = grades[num]

wb.save('grades_v1.xlsx')

        
